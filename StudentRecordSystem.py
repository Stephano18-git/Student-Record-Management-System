import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

file_name = "student_records.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(file_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(["Student ID", "Name", "Course", "GPA", "Timestamp"])
    workbook.save(file_name)
    print("✅ Excel file created successfully!")
else:
    print("📁 Excel file already exists.")


# ADD NEW RECORD
def add_record():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    student_id = input("Enter Student ID: ")
    name = input("Enter Name: ")
    course = input("Enter Course: ")

    # GPA validation
    while True:
        try:
            gpa = float(input("Enter GPA (0.0 - 4.0): "))
            if 0.0 <= gpa <= 4.0:
                break
            else:
                print("GPA must be between 0.0 and 4.0. Please try again.")
        except ValueError:
            print("Please enter a valid number.")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([student_id, name, course, gpa, timestamp])
    workbook.save(file_name)
    print(f"✅ Record for {name} added successfully!")


# VIEW ALL RECORDS
def view_records():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    print("\nAll Student Records:\n")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)


# EDIT RECORD
def edit_record():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    student_id = input("Enter Student ID to edit: ")
    found = False

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == student_id:
            found = True
            row[1].value = input("Enter new Name: ")
            row[2].value = input("Enter new Course: ")

            while True:
                try:
                    gpa = float(input("Enter new GPA (0.0 - 4.0): "))
                    if 0.0 <= gpa <= 4.0:
                        row[3].value = gpa
                        break
                    else:
                        print("GPA must be between 0.0 and 4.0.")
                except ValueError:
                    print("Please enter a valid number.")

            row[4].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print("✅ Record updated successfully!")
            break

    if not found:
        print("❌ Student ID not found.")

    workbook.save(file_name)


# DELETE RECORD
def delete_record():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    student_id = input("Enter Student ID to delete: ")
    found = False

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == student_id:
            sheet.delete_rows(row)
            found = True
            print("🗑️ Record deleted successfully!")
            break

    if not found:
        print("❌ Student ID not found.")

    workbook.save(file_name)


# MENU SYSTEM
while True:
    print("\n" + "=" * 50)
    print("STUDENT RECORD MANAGEMENT SYSTEM")
    print("=" * 50)
    print("1. Add New Record")
    print("2. View All Records")
    print("3. Edit Record")
    print("4. Delete Record")
    print("5. Exit")

    choice = input("Enter your choice (1-5): ")

    if choice == '1':
        add_record()
    elif choice == '2':
        view_records()
    elif choice == '3':
        edit_record()
    elif choice == '4':
        delete_record()
    elif choice == '5':
        print("Exiting program... Goodbye!")
        break
    else:
        print("Invalid choice. Please try again.")
